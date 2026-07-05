import csv
import io

from fastapi import APIRouter, Depends, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from starlette.responses import StreamingResponse

from app.database import get_db
from app.dependencies import get_current_user
from app.models.property import Property
from app.models.user import User

router = APIRouter(prefix="/api/export", tags=["export"])


def _build_csv_data(rows):
    output = io.StringIO()
    writer = csv.writer(output)

    extra_keys: list[str] = []
    for p in rows:
        if p.extra_data:
            for k in p.extra_data:
                if k not in extra_keys:
                    extra_keys.append(k)

    headers = ["ID", "Наименование", "Адрес", "Нормализованный адрес", "Ссылка", "Пользователь", "Дата добавления"] + extra_keys
    writer.writerow(headers)

    for p in rows:
        row = [
            p.id,
            p.name,
            p.address,
            p.normalized_address,
            p.link or "",
            p.user.nickname if p.user else "",
            p.created_at.isoformat(),
        ]
        for k in extra_keys:
            row.append(p.extra_data.get(k, "") if p.extra_data else "")
        writer.writerow(row)

    output.seek(0)
    return output.getvalue()


@router.get("/properties/csv")
async def export_properties_csv(
    current_user: User | None = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Property)
        .options(selectinload(Property.user))
        .order_by(Property.created_at.desc())
    )
    rows = result.scalars().all()

    data = _build_csv_data(rows)

    return StreamingResponse(
        iter([data]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=properties.csv"},
    )


@router.get("/properties/xlsx")
async def export_properties_xlsx(
    current_user: User | None = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=501,
            content={"detail": "Установите openpyxl: pip install openpyxl"}
        )

    result = await db.execute(
        select(Property)
        .options(selectinload(Property.user))
        .order_by(Property.created_at.desc())
    )
    rows = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Объекты недвижимости"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    extra_keys: list[str] = []
    for p in rows:
        if p.extra_data:
            for k in p.extra_data:
                if k not in extra_keys:
                    extra_keys.append(k)

    headers = ["ID", "Наименование", "Адрес", "Нормализованный адрес", "Ссылка", "Пользователь", "Дата добавления"] + extra_keys
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, p in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=p.id)
        ws.cell(row=row_idx, column=2, value=p.name)
        ws.cell(row=row_idx, column=3, value=p.address)
        ws.cell(row=row_idx, column=4, value=p.normalized_address)
        ws.cell(row=row_idx, column=5, value=p.link or "")
        ws.cell(row=row_idx, column=6, value=p.user.nickname if p.user else "")
        ws.cell(row=row_idx, column=7, value=p.created_at.isoformat())
        for ek_idx, k in enumerate(extra_keys):
            ws.cell(row=row_idx, column=8 + ek_idx, value=p.extra_data.get(k, "") if p.extra_data else "")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=properties.xlsx"},
    )
